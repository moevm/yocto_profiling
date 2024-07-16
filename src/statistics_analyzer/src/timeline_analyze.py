import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, PatternFill, Font


def get_tasks(parser, border, resource):
    tasks = []
    newline = False
    for time, info in parser.timeline.items():
        if not info[resource] is None and info[resource] < border:
            newline = True
            for package in parser.info:
                for task in parser.info[package]:
                    if time > float(parser.info[package][task]['Started']) and time < float(parser.info[package][task]['Ended']):
                        tasks.append(f'{time}: {package}.{task}\n')
        else:
            if newline:
                tasks.append('\n')
                newline = False
    return tasks

def get_tasks_for_intervals(parser, intervals):
    for interval in intervals:
        for time in range(interval[0], interval[1]):
            for package in parser.info:
                for task in parser.info[package]:
                    if time > float(parser.info[package][task]['Started']) and time < float(parser.info[package][task]['Ended']):
                        interval[3].add(f'{package}.{task}')
    return intervals


def find_free_intervals(parser, resource, border):
    intervals = []
    is_free = False
    first_timestamp = None
    sum_time = 0
    for time, info in parser.timeline.items():
        if not info[resource] is None and info[resource] < border:
            if not is_free:
                is_free = True
                first_timestamp = time
        else:
            if info[resource] and info[resource] > border:
                if is_free:
                    is_free = False
                    intervals.append([first_timestamp, time, time - first_timestamp, set()])
                    sum_time += time - first_timestamp
    return sorted(intervals, key=lambda x: x[0]), sum_time


def write_to_json(intervals, sum_time, filename):
    items = {'sum_time': sum_time, 'items': []}
    for interval in intervals:
        items['items'].append({'Start': interval[0], 'End': interval[1], 'Duration': interval[2], 'tasks': list(interval[3])})

    with open(filename, 'w') as file:
        json.dump(items, file, indent=4)



def write_to_excel(parser):
    wb = Workbook()
    for sheet in wb.sheetnames:
        wb.remove(wb[sheet])
    ws = wb.create_sheet('sources')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 125
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 50

    first_timestamp = 0
    ws.cell(row=1, column=1).value = 'Sec'
    ws.cell(row=1, column=2).value = 'CPU'
    ws.cell(row=1, column=3).value = 'IO'
    ws.cell(row=1, column=4).value = 'RAM'
    ws.cell(row=1, column=5).value = 'Running tasks'
    ws.cell(row=1, column=6).value = 'Buildable tasks'
    ws.cell(row=1, column=7).value = 'Buildable task types'
    ws.cell(row=1, column=8).value = 'Skip start running info'

    for i in range(1, 9):
        ws.cell(row=1, column=i).font = Font(bold=True)

    for index, time in enumerate(dict(sorted(parser.timeline.items()))):
        if not first_timestamp:
            first_timestamp = time
        ws.cell(row=index+2, column=1).value = time - first_timestamp

        if not parser.timeline[time]['cpu'] is None:
            ws.cell(row=index+2, column=2).value = str(round(parser.timeline[time]['cpu'], 4))
            cpu_color = Color(
                rgb=f'ff{hex(int(255 * parser.timeline[time]["cpu"]))[2:].zfill(2)}{hex(int(255 * (1 - parser.timeline[time]["cpu"])))[2:].zfill(2)}{hex(0)[2:].zfill(2)}')
            ws.cell(row=index+2, column=2).fill = PatternFill(fgColor=cpu_color, fill_type='solid')
        else:
            ws.cell(row=index+2, column=2).fill = PatternFill(fgColor=Color(rgb='ffffffff'), fill_type='solid')

        if not parser.timeline[time]['io'] is None:
            ws.cell(row=index+2, column=3).value = str(round(parser.timeline[time]['io'], 4))
            io_color = Color(
                rgb=f'ff{hex(int(255 * parser.timeline[time]["io"]))[2:].zfill(2)}{hex(int(255 * (1 - parser.timeline[time]["io"])))[2:].zfill(2)}{hex(0)[2:].zfill(2)}')
            ws.cell(row=index+2, column=3).fill = PatternFill(fgColor=io_color, fill_type='solid')
        else:
            ws.cell(row=index+2, column=3).fill = PatternFill(fgColor=Color(rgb='ffffffff'), fill_type='solid')

        if not parser.timeline[time]['ram'] is None:
            ws.cell(row=index+2, column=4).value = str(round(parser.timeline[time]['ram'], 4))
            ram_color = Color(
                rgb=f'ff{hex(int(255 * parser.timeline[time]["ram"]))[2:].zfill(2)}{hex(int(255 * (1 - parser.timeline[time]["ram"])))[2:].zfill(2)}{hex(0)[2:].zfill(2)}')
            ws.cell(row=index+2, column=4).fill = PatternFill(fgColor=ram_color, fill_type='solid')
        else:
            ws.cell(row=index+2, column=4).fill = PatternFill(fgColor=Color(rgb='ffffffff'), fill_type='solid')

        ws.cell(row=index+2, column=5).value = '\n'.join(parser.timeline[time]['tasks'])

        if time in parser.queue:
            ws.cell(row=index+2, column=6).value = '\n'.join(parser.queue[time]['tasks'])
            ws.cell(row=index+2, column=7).value = '\n'.join([f'{key}: {value}' for key, value in parser.queue[time]['task_types'].items()])

        if time in parser.skipped_info:
            ws.cell(row=index+2, column=8).value = '\n'.join(parser.skipped_info[time])

        for i in range(1, 9):
            ws.cell(row=index+2, column=i).alignment = Alignment(wrap_text=True)

    wb.save('./src/statistics_analyzer/output/sources.xlsx')